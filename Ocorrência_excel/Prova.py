import openpyxl
workbook = openpyxl.load_workbook('ocorrencia.xlsx')
planilha = workbook['cadastros']
planilha = workbook.active
import os # = importa o limpa tela --> os.system('cls')
import time # = importa tempo de tela --> time.sleep()
from tabulate import tabulate 

def cadastro():
    print("=======================MENU DE CADASTRO DE OCORRÊNCIA======================="
        "\n"
        )
    nome = input("INFORME O NOME DO SOLICITANTE: ")
    cpf = input("INFORME O CPF: ")
    ende = input("INFORME O ENDEREÇO: ")
    tel = input("INFORME O TELEFONE: ")
    hor = input("INFORME O HORÁRIO: ")
    desc_ocorrido = input("DESCRIÇÃO DO OCORRIDO: ")
    obs = input("OBSERVAÇÃO: ")
    os.system('cls')
    print("SALVANDO CADASTRO...")
    time.sleep(3)
    os.system('cls')
    nova_linha = [nome,cpf,ende,tel,hor,desc_ocorrido,obs]
    planilha.append(nova_linha)
    workbook.save('ocorrencia.xlsx')

def listagem():
    print("                                                           =======================LISTAGEM DE OCORRÊNCIA======================="
        "\n"
        )
    #PARA VER A LISTAGEM COMPLETA, DEIXAR O TERMINAL EM TELA CHEIA.
    armazenamento = [
        'NOME',
        'CPF',
        'ENDEREÇO',
        'TELEFONE',
        'HORÁRIO',
        'DESCRIÇÃO DO OCORRIDO',
        'OBSERVAÇÃO'
    ]
    dados = []
    i = 0
    for row in planilha.iter_rows(min_row=3, max_col=7,):
        dados.append([])
        for cell in row:
            dados[i].append(cell.value)
        i = i + 1
    print(tabulate(dados, headers= armazenamento, tablefmt="rounded_grid", stralign="center", numalign="center"))

def listagem_1():
    dados = []
    for row in planilha.iter_rows(values_only=True):
        dados.append(row)
    tabela_formatada = tabulate(dados, tablefmt="rounded_grid", stralign="center");
    print(tabela_formatada)

def listagem_1():
    dados = []
    for row in planilha.iter_rows(values_only=True):
        dados.append(row)
    tabela_formatada = tabulate(dados, tablefmt="rounded_grid", stralign="center");
    print(tabela_formatada)

def atualizar():
    print("=======================ATUALIZAÇÃO DE OCORRÊNCIA=======================")
    print("")
    att_cpf = input("INFORME O CPF DA OCORRÊNCIA QUE DESEJA ATUALIZAR --> ")
    for row in planilha.iter_rows(min_row=2, max_col=7,):
        for cell in row:
            if (str(cell.value) == att_cpf):
                linha = cell.row
                print("BUSCANDO...")
                time.sleep(3)
                os.system('cls')
                nova_obs = input("INFORME A NOVA OBSERVAÇÃO: ")
                for i in range(1,7):
                    planilha.cell(linha, 7, nova_obs)
                    print("ATUALIZANDO...")
                    time.sleep(3)
                    os.system('cls')
                    workbook.save('ocorrencia.xlsx')

while True:
    print(
        "                     PREFEITURA MUNICIPAL DE JUIZ DE FORA                     "
        "\n                                    PMJF                     "
        "\n"
    )
    
    print("BEM VINDO AO SISTEMA DE ARMAZENAMENTO DE OCORRÊNCIAS EM APOIO COM A DEFESA CIVIL DE JUIZ DE FORA")
    print(
        "\n!O SISTEMA DE OCORRÊNCIA BUSCA AUXILIAR E AJUDAR AS VÍTIMAS DO TEMPORAL NA CIDADE!"
        "\n|VEJA ABAIXO AS OPÇÕES DISPONÍVEIS|"
        )
    print(
        "\n[1] CADASTRAR"
        "\n[2] LISTAR"
        "\n[3] ATUALIZAR"
        "\n[0] SAIR DO SISTEMA"
        "\n "
    )
    op = input("INFORME A OPÇÃO DESEJADA --> ")

    if op == "1":
        os.system('cls')
        cadastro()
    if op == "2":
        listagem_1()
    if op == "3":
        os.system('cls')
        atualizar()
    if op == "0":
        os.system('cls')
        print("SAINDO DO SISTEMA...")
        time.sleep(3)
        break
